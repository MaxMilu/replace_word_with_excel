import shutil
from time import sleep
from docxtpl import DocxTemplate
from openpyxl import load_workbook
import os

# 配置路径和文件名
BASE_DIR = "./work"  # 基础路径
OUTPUT_DIR = "./output"  # 输出路径
OUTPUT_FILE_DIRS = ["房主-身份证", "车主-身份证"]  # 输出文件夹名称
OUTPUT_FILE_SUFFIX = ".docx"  # 输出文件后缀
DATA_FILE = BASE_DIR + "/数据.xlsx"  # 数据文件
DATA_FILE_SHEETS = ["房主-身份证", "车主-身份证"]  # 数据文件中的 sheet名称
TEMPLATE_FILE = BASE_DIR + "/模板.docx"  # 模板文件


# 根据路径创建文件夹
def create_path(path):
    if not os.path.exists(path):
        os.makedirs(path)


# 创建类别文件夹
def create_need_dirs():
    # base dir
    create_path(os.path.join(BASE_DIR))
    create_path(os.path.join(OUTPUT_DIR))

    for category in OUTPUT_FILE_DIRS:
        create_path(os.path.join(OUTPUT_DIR, category))


def delete_all_files_in_directory(directory):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.remove(file_path)
        elif os.path.isdir(file_path):
            shutil.rmtree(file_path)


# 读取Excel数据
def read_excel_data(file_path):
    start_row = 2
    wb = load_workbook(file_path)
    contexts = []
    for currentSheetName in DATA_FILE_SHEETS:
        ws = wb.get_sheet_by_name(currentSheetName)
        for row in range(start_row, ws.max_row + 1):
            context = {
                "sheetName": str(ws.title),
                "no": str(ws[f"A{row}"].value).strip(),  # 序号 A
                "name": str(ws[f"B{row}"].value).strip(),  # 姓名 B
                "id": str(ws[f"C{row}"].value).strip(),  # 身份证号 C
                "location": str(ws[f"D{row}"].value).strip(),  # 住址号 D
                "phone": str(ws[f"E{row}"].value).strip(),  # 电话 E
                "startDate": str(ws[f"S{row}"].value).strip(),  # 起始日期 S
                "endDate": str(ws[f"T{row}"].value).strip(),  # 结束日期 T
                "houseArea": str(ws[f"I{row}"].value).strip(),  # 建筑面积 I
                "amount": str(ws[f"J{row}"].value).strip(),  # 欠费金额 J
                # "liquidatedDamages": str(ws[f"A{row}"].value).strip(),  # 违约金 K
            }
            # 判断有序号的情况下才添加
            if context["no"] != 'None' and context["no"].isdigit():
                contexts.append(context)
    return contexts


# 生成Word文档
def generate_word_documents(contexts, template_path):
    for context in contexts:
        tpl = DocxTemplate(template_path)
        tpl.render(context)
        output_path = os.path.join(OUTPUT_DIR, context["sheetName"])
        person_file_name = context["name"]
        no = context["no"]
        output_file = os.path.join(output_path, f"{no}.{person_file_name}{OUTPUT_FILE_SUFFIX}")
        tpl.save(output_file)
        print(f"生成文件: {output_file}")
        sleep(1)  # 间隔1秒,为了方便按照创建时间排序


def main():
    delete_all_files_in_directory(OUTPUT_DIR)
    create_need_dirs()
    contexts = read_excel_data(DATA_FILE)
    generate_word_documents(contexts, TEMPLATE_FILE)
    print("处理完成")


if __name__ == "__main__":
    main()
